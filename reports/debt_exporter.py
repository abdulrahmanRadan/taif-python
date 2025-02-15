import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from database.database_manager import DatabaseManager

class DebtExporter:
    def __init__(self, master, debt_service):
        self.master = master
        self.debt_service = debt_service
        self.db_manager = DatabaseManager()

        # إنشاء نافذة التصدير
        self.export_window = tk.Toplevel(self.master)
        self.export_window.title("تصدير بيانات الديون")
        self.export_window.geometry("400x200")
        self.export_window.grab_set()
        self.center_window()

        # المتغيرات
        self.filename = tk.StringVar(value="تصدير_الديون")

        # إنشاء الواجهة
        self.create_widgets()

    def center_window(self):
        self.export_window.update_idletasks()
        width = self.export_window.winfo_width()
        height = self.export_window.winfo_height()
        x = (self.export_window.winfo_screenwidth() // 2) - (width // 2)
        y = (self.export_window.winfo_screenheight() // 2) - (height // 2)
        self.export_window.geometry(f"{width}x{height}+{x}+{y}")

    def create_widgets(self):
        form_frame = ttk.Frame(self.export_window, padding=10)
        form_frame.pack(fill=tk.BOTH, expand=True)

        # حقل اسم الملف
        ttk.Label(form_frame, text="اسم الملف:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        filename_entry = ttk.Entry(form_frame, textvariable=self.filename, width=30)
        filename_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # زر التصدير
        export_btn = ttk.Button(
            form_frame,
            text="تصدير إلى Excel",
            command=self.export_to_excel
        )
        export_btn.grid(row=1, column=0, columnspan=2, pady=10, sticky="ew")

    def format_currency(self, value, currency_code):
        """تنسيق العملة بناءً على نوع الخدمة"""
        currency_map = {
            "1": "ر.ي",
            "2": "ر.س",
            "3": "دولار"
        }
        return f"{value} {currency_map.get(currency_code, 'ر.ي')}"

    def apply_rtl_to_excel(self, file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        
        # تطبيق محاذاة النص
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # ضبط عرض الأعمدة تلقائيًا
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        wb.save(file_path)

    def apply_colors_to_excel(self, file_path):
        wb = load_workbook(file_path)
        ws = wb.active

        # تعريف الأنماط
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        even_row = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        odd_row = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        warning_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # تلوين الرأس
        for cell in ws[1]:
            cell.fill = header_fill

        # تلوين الصفوف
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = even_row if row_idx % 2 == 0 else odd_row
            for cell in row:
                cell.fill = fill

        # تلوين الخلايا التي بها ديون
        for row in ws.iter_rows(min_row=2):
            remaining_cell = row[5]  # عمود المبلغ المتبقي
            try:
                if float(remaining_cell.value.split()[0]) > 0:
                    remaining_cell.fill = warning_fill
            except:
                pass

        wb.save(file_path)

    def export_to_excel(self):
        try:
            # جلب بيانات الديون
            debts = self.debt_service.get_all_data()
            
            # تحويل إلى DataFrame
            df = pd.DataFrame(debts)
            
            # إعادة تسمية الأعمدة
            df.rename(columns={
                "id": "الرقم",
                "name": "الاسم",
                "type": "نوع الخدمة",
                "date": "التاريخ",
                "ym_paid": "المبلغ المدفوع (يمني)",
                "sm_paid": "المبلغ المدفوع (سعودي)",
                "remaining": "المتبقي"
            }, inplace=True)

            # جلب بيانات المدفوعات لكل دين
            payments_data = []
            for debt in debts:
                payments = self.debt_service.get_payments(debt["type"], debt["id"])
                if payments:
                    for payment in payments:
                        payments_data.append({
                            "الرقم": debt["id"],
                            "الدفعة": f"{payment['amount']} ",
                            "تاريخ الدفعة": payment["payment_date"],
                            "طريقة الدفع": payment["payment_method"]
                        })

            # تحويل بيانات المدفوعات إلى DataFrame
            if payments_data:
                payments_df = pd.DataFrame(payments_data)
                # دمج بيانات الديون والمدفوعات
                df = pd.merge(df, payments_df, on="الرقم", how="left")

            # حفظ الملف
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            file_path = os.path.join(downloads_path, f"{self.filename.get()}.xlsx")
            df.to_excel(file_path, index=False)

            # تطبيق التنسيقات
            self.apply_rtl_to_excel(file_path)
            self.apply_colors_to_excel(file_path)

            messagebox.showinfo("نجاح", f"تم التصدير بنجاح إلى:\n{file_path}")
            self.export_window.destroy()

        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء التصدير: {str(e)}")