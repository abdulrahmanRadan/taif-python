import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import pandas as pd
import os
from database.database_manager import DatabaseManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

class PassportsExporter:
    def __init__(self, master):
        self.master = master
        self.table_name = "Passports"  # اسم الجدول
        self.db_manager = DatabaseManager()

        # إنشاء نافذة التصدير
        self.export_window = tk.Toplevel(master)
        self.export_window.title("تصدير بيانات الجوازات")
        self.export_window.geometry("400x400")  # زيادة الارتفاع لإضافة الحقول الجديدة
        self.export_window.grab_set()  # جعل النافذة إجبارية
        self.center_window()

        # المتغيرات
        self.filename = tk.StringVar(value="تصدير_الجوازات")
        self.export_option = tk.StringVar(value="جميع البيانات")
        self.selected_date = tk.StringVar()  # متغير لتخزين التاريخ المحدد
        self.remaining_amount_threshold = tk.DoubleVar(value=0.0)  # متغير جديد لتخزين قيمة "المبلغ المتبقي"
        self.passport_type = tk.StringVar()  # متغير لنوع الجواز
        self.passport_status = tk.StringVar()  # متغير لحالة الجواز

        # إنشاء الواجهة
        self.create_widgets()

    def center_window(self):
        """
        وضع النافذة في وسط الشاشة.
        """
        self.export_window.update_idletasks()
        width = self.export_window.winfo_width()
        height = self.export_window.winfo_height()
        x = (self.export_window.winfo_screenwidth() // 2) - (width // 2)
        y = (self.export_window.winfo_screenheight() // 2) - (height // 2)
        self.export_window.geometry(f"{width}x{height}+{x}+{y}")

    def create_widgets(self):
        """
        إنشاء واجهة المستخدم.
        """
        form_frame = ttk.Frame(self.export_window, padding=10)
        form_frame.pack(fill=tk.BOTH, expand=True)

        # حقل اسم الملف
        ttk.Label(form_frame, text="اسم الملف:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        filename_entry = ttk.Entry(form_frame, textvariable=self.filename, width=30)
        filename_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # قائمة منسدلة لخيارات التصدير
        ttk.Label(form_frame, text="خيارات التصدير:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        export_options = ttk.Combobox(
            form_frame,
            textvariable=self.export_option,
            values=["جميع البيانات", "حسب التاريخ", "بيانات بها مبالغ متبقية", "حسب نوع الجواز", "حسب حالة الجواز"],
            state="readonly"
        )
        export_options.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        export_options.bind("<<ComboboxSelected>>", self.toggle_fields)

        # حقل التاريخ (مخفي بشكل افتراضي)
        self.date_label = ttk.Label(form_frame, text="التاريخ:")
        self.date_entry = DateEntry(form_frame, textvariable=self.selected_date, date_pattern="yyyy-mm-dd")
        self.date_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.date_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.date_label.grid_remove()
        self.date_entry.grid_remove()

        # حقل "المبلغ المتبقي" (مخفي بشكل افتراضي)
        self.remaining_amount_label = ttk.Label(form_frame, text="المبلغ المتبقي أقل من:")
        self.remaining_amount_entry = ttk.Entry(form_frame, textvariable=self.remaining_amount_threshold, width=30)
        self.remaining_amount_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")
        self.remaining_amount_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        self.remaining_amount_label.grid_remove()
        self.remaining_amount_entry.grid_remove()

        # قائمة منسدلة لأنواع الجوازات (مخفية بشكل افتراضي)
        self.passport_type_label = ttk.Label(form_frame, text="نوع الجواز:")
        self.passport_type_combobox = ttk.Combobox(form_frame, textvariable=self.passport_type, state="readonly")
        self.passport_type_combobox.grid(row=4, column=1, padx=5, pady=5, sticky="w")
        self.passport_type_combobox.grid_remove()  # إخفاء القائمة المنسدلة افتراضيًا

        # قائمة منسدلة لحالات الجوازات (مخفية بشكل افتراضي)
        self.passport_status_label = ttk.Label(form_frame, text="حالة الجواز:")
        self.passport_status_combobox = ttk.Combobox(form_frame, textvariable=self.passport_status, state="readonly")
        self.passport_status_combobox.grid(row=5, column=1, padx=5, pady=5, sticky="w")
        self.passport_status_combobox.grid_remove()  # إخفاء القائمة المنسدلة افتراضيًا

        # زر التصدير إلى Excel
        export_excel_button = ttk.Button(form_frame, text="تصدير إلى Excel", command=self.export_to_excel)
        export_excel_button.grid(row=6, column=0, columnspan=2, pady=10, padx=5, sticky="ew")

    def toggle_fields(self, event=None):
        """
        إظهار أو إخفاء حقول التاريخ وحقل "المبلغ المتبقي" بناءً على الخيار المحدد.
        """
        if self.export_option.get() == "حسب التاريخ":
            self.date_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
            self.date_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
            self.remaining_amount_label.grid_remove()
            self.remaining_amount_entry.grid_remove()
            self.passport_type_label.grid_remove()
            self.passport_type_combobox.grid_remove()
            self.passport_status_label.grid_remove()
            self.passport_status_combobox.grid_remove()
        elif self.export_option.get() == "بيانات بها مبالغ متبقية":
            self.remaining_amount_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")
            self.remaining_amount_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")
            self.date_label.grid_remove()
            self.date_entry.grid_remove()
            self.passport_type_label.grid_remove()
            self.passport_type_combobox.grid_remove()
            self.passport_status_label.grid_remove()
            self.passport_status_combobox.grid_remove()
        elif self.export_option.get() == "حسب نوع الجواز":
            self.passport_type_label.grid(row=4, column=0, padx=5, pady=5, sticky="e")
            self.passport_type_combobox.grid(row=4, column=1, padx=5, pady=5, sticky="w")
            self.load_passport_types()  # تحميل أنواع الجوازات
            self.date_label.grid_remove()
            self.date_entry.grid_remove()
            self.remaining_amount_label.grid_remove()
            self.remaining_amount_entry.grid_remove()
            self.passport_status_label.grid_remove()
            self.passport_status_combobox.grid_remove()
        elif self.export_option.get() == "حسب حالة الجواز":
            self.passport_status_label.grid(row=5, column=0, padx=5, pady=5, sticky="e")
            self.passport_status_combobox.grid(row=5, column=1, padx=5, pady=5, sticky="w")
            self.load_passport_statuses()  # تحميل حالات الجوازات
            self.date_label.grid_remove()
            self.date_entry.grid_remove()
            self.remaining_amount_label.grid_remove()
            self.remaining_amount_entry.grid_remove()
            self.passport_type_label.grid_remove()
            self.passport_type_combobox.grid_remove()
        else:
            self.date_label.grid_remove()
            self.date_entry.grid_remove()
            self.remaining_amount_label.grid_remove()
            self.remaining_amount_entry.grid_remove()
            self.passport_type_label.grid_remove()
            self.passport_type_combobox.grid_remove()
            self.passport_status_label.grid_remove()
            self.passport_status_combobox.grid_remove()

    def load_passport_types(self):
        """
        تحميل أنواع الجوازات من قاعدة البيانات.
        """
        query = f"SELECT DISTINCT type FROM {self.table_name}"
        types = self.db_manager.execute_read_query(query)
        if types:
            type_codes = [str(type[0]) for type in types]
            type_names = [self.format_type(code) for code in type_codes]
            self.passport_type_combobox["values"] = type_names

    def load_passport_statuses(self):
        """
        تحميل حالات الجوازات من قاعدة البيانات.
        """
        query = f"SELECT DISTINCT status FROM {self.table_name}"
        statuses = self.db_manager.execute_read_query(query)
        if statuses:
            status_codes = [str(status[0]) for status in statuses]
            status_names = [self.format_status(code) for code in status_codes]
            self.passport_status_combobox["values"] = status_names

    def format_type(self, type_code):
        """
        تحويل رمز نوع الجواز المخزن في قاعدة البيانات إلى نص.
        """
        type_map = {"1": "عادي", "2": "مستعجل عدن", "3": "مستعجل بيومه", "4": "غير ذلك"}
        return type_map.get(type_code, "غير معروف")

    def format_status(self, status_code):
        """
        تحويل رمز حالة الجواز المخزن في قاعدة البيانات إلى نص.
        """
        status_map = {"1": "في الطابعة", "2": "في المكتب", "3": "تم الاستلام", "4": "مرفوض"}
        return status_map.get(status_code, "غير معروف")

    def get_filtered_data(self):
        """
        جلب البيانات المصفاة بناءً على الخيارات المحددة.
        """
        query = f"SELECT * FROM {self.table_name}"
        conditions = []

        if self.export_option.get() == "حسب التاريخ":
            if not self.selected_date.get():
                messagebox.showwarning("تحذير", "يجب تحديد التاريخ.")
                return None
            conditions.append(f"(booking_date = '{self.selected_date.get()}' OR receipt_date = '{self.selected_date.get()}')")

        elif self.export_option.get() == "بيانات بها مبالغ متبقية":
            conditions.append(f"remaining_amount <= {self.remaining_amount_threshold.get()}")

        elif self.export_option.get() == "حسب نوع الجواز":
            if not self.passport_type.get():
                messagebox.showwarning("تحذير", "يجب تحديد نوع الجواز.")
                return None
            type_code = self.get_type_code(self.passport_type.get())
            conditions.append(f"type = '{type_code}'")

        elif self.export_option.get() == "حسب حالة الجواز":
            if not self.passport_status.get():
                messagebox.showwarning("تحذير", "يجب تحديد حالة الجواز.")
                return None
            status_code = self.get_status_code(self.passport_status.get())
            conditions.append(f"status = '{status_code}'")

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        # جلب البيانات من قاعدة البيانات
        data = self.db_manager.execute_read_query(query)

        if not data:
            messagebox.showinfo("معلومات", "لا توجد بيانات للتصدير.")
            return None

        return data

    def get_type_code(self, type_name):
        """
        تحويل اسم نوع الجواز إلى رمزه.
        """
        type_map = {"عادي": "1", "مستعجل عدن": "2", "مستعجل بيومه": "3", "غير ذلك": "4"}
        return type_map.get(type_name, "4")

    def get_status_code(self, status_name):
        """
        تحويل اسم حالة الجواز إلى رمزه.
        """
        status_map = {"في الطابعة": "1", "في المكتب": "2", "تم الاستلام": "3", "مرفوض": "4"}
        return status_map.get(status_name, "1")

    def format_currency(self, currency_code):
        """
        تحويل رمز العملة المخزن في قاعدة البيانات إلى نص.
        """
        currency_map = {"1": "ر.ي", "2": "ر.س", "3": "دولار"}
        return currency_map.get(str(currency_code), "ر.ي")  # افتراضيًا ر.ي إذا لم يتم العثور على الرمز

    def merge_and_remove_currency(self, df):
        """
        دمج العملة مع الأعمدة المالية وحذف عمود العملة.
        """
        # تحويل رمز العملة إلى نص باستخدام دالة format_currency
        df["العملة"] = df["العملة"].apply(self.format_currency)

        # دمج العملة مع الأعمدة المالية
        df["سعر الحجز"] = df["سعر الحجز"].astype(str) + " " + df["العملة"]
        df["سعر الشراء"] = df["سعر الشراء"].astype(str) + " " + df["العملة"]
        df["المبلغ الصافي"] = df["المبلغ الصافي"].astype(str) + " " + df["العملة"]
        df["المبلغ المدفوع"] = df["المبلغ المدفوع"].astype(str) + " " + df["العملة"]
        df["المبلغ المتبقي"] = df["المبلغ المتبقي"].astype(str) + " " + df["العملة"]

        # حذف عمود العملة
        df.drop(columns=["العملة"], inplace=True)

        return df

    def apply_rtl_to_excel(self, file_path):
        """
        تطبيق توجيه النص من اليمين إلى اليسار على ملف Excel.
        """
        # تحميل ملف Excel
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # تطبيق توجيه النص RTL على جميع الخلايا
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="right", vertical="center")

        # حفظ التغييرات
        workbook.save(file_path)

    def apply_colors_to_excel(self, file_path):
        """
        تطبيق الألوان على الصفوف والأعمدة في ملف Excel.
        """
        # تحميل ملف Excel
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # تعريف الألوان
        gray_100 = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # رمادي فاتح
        white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # أبيض
        orange_light = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # برتقالي فاتح

        # تطبيق الألوان على الصفوف
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # تجاهل الصف الأول (العناوين)
            if row_idx % 2 == 0:  # الصفوف الزوجية
                for cell in row:
                    cell.fill = gray_100
            else:  # الصفوف الفردية
                for cell in row:
                    cell.fill = white

        # تطبيق ألوان خاصة على أعمدة معينة
        for col_idx, col in enumerate(sheet.iter_cols(min_row=1, max_row=1), start=1):  # العناوين
            for cell in col:
                if cell.value in ["سعر الحجز", "سعر الشراء", "المبلغ الصافي", "المبلغ المدفوع", "المبلغ المتبقي"]:
                    for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  # برتقالي فاتح

        # حفظ التغييرات
        workbook.save(file_path)

    def export_to_excel(self):
        """
        تصدير البيانات إلى ملف Excel.
        """
        data = self.get_filtered_data()
        if not data:
            return

        # تحويل البيانات إلى DataFrame
        columns = [description[0] for description in self.db_manager.cursor.description]
        df = pd.DataFrame(data, columns=columns)

        # تحويل أسماء الأعمدة إلى العربية
        arabic_columns = {
            "id": "الرقم",
            "name": "الاسم",
            "booking_date": "تاريخ الحجز",
            "type": "النوع",
            "booking_price": "سعر الحجز",
            "purchase_price": "سعر الشراء",
            "net_amount": "المبلغ الصافي",
            "paid_amount": "المبلغ المدفوع",
            "remaining_amount": "المبلغ المتبقي",
            "status": "الحالة",
            "receipt_date": "تاريخ الاستلام",
            "receiver_name": "اسم المستلم",
            "currency": "العملة"
        }
        df.rename(columns=arabic_columns, inplace=True)

        # دمج العملة مع الأعمدة المالية وحذف عمود العملة
        df = self.merge_and_remove_currency(df)

        # تحديد مسار حفظ الملف
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        file_path = os.path.join(downloads_path, f"{self.filename.get()}.xlsx")

        # تصدير البيانات إلى Excel
        df.to_excel(file_path, index=False)

        # تطبيق توجيه النص RTL على ملف Excel
        self.apply_rtl_to_excel(file_path)

        # تطبيق الألوان على ملف Excel
        self.apply_colors_to_excel(file_path)

        messagebox.showinfo("نجاح", f"تم تصدير البيانات بنجاح إلى: {file_path}")

        # إغلاق النافذة بعد التصدير
        self.export_window.destroy()