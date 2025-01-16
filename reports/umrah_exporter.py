import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import pandas as pd
import os
from database.database_manager import DatabaseManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

class UmrahExporter:
    def __init__(self, master):
        self.master = master
        self.table_name = "Umrah"  # اسم الجدول
        self.db_manager = DatabaseManager()

        # إنشاء نافذة التصدير
        self.export_window = tk.Toplevel(master)
        self.export_window.title("تصدير بيانات العمرة")
        self.export_window.geometry("400x300")  # زيادة الارتفاع لإضافة الحقل الجديد
        self.export_window.grab_set()  # جعل النافذة إجبارية
        self.center_window()

        # المتغيرات
        self.filename = tk.StringVar(value="تصدير_العمرة")
        self.export_option = tk.StringVar(value="جميع البيانات")
        self.entry_date = tk.StringVar()
        self.exit_date = tk.StringVar()
        self.remaining_amount_threshold = tk.DoubleVar(value=0.0)  # متغير جديد لتخزين قيمة "المبلغ المتبقي"

        # إنشاء الواجهة
        self.create_widgets()

    def center_window(self):
        """
        وضع النافذة في وسط الشاشة.
        """
        self.export_window.update_idletasks()
        width = self.export_window.winfo_width()
        height = self.export_window.winfo_height()
        x = (self.export_window.winfo_screenwidth() // 2) - (width // 2)
        y = (self.export_window.winfo_screenheight() // 2) - (height // 2)
        self.export_window.geometry(f"{width}x{height}+{x}+{y}")

    def create_widgets(self):
        """
        إنشاء واجهة المستخدم.
        """
        form_frame = ttk.Frame(self.export_window, padding=10)
        form_frame.pack(fill=tk.BOTH, expand=True)

        # حقل اسم الملف
        ttk.Label(form_frame, text="اسم الملف:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        filename_entry = ttk.Entry(form_frame, textvariable=self.filename, width=30)
        filename_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # قائمة منسدلة لخيارات التصدير
        ttk.Label(form_frame, text="خيارات التصدير:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        export_options = ttk.Combobox(
            form_frame,
            textvariable=self.export_option,
            values=["جميع البيانات", "حسب تاريخ الدخول", "حسب تاريخ الخروج", "بيانات بها مبالغ متبقية"],
            state="readonly"
        )
        export_options.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        export_options.bind("<<ComboboxSelected>>", self.toggle_fields)

        # حقول التاريخ (مخفية بشكل افتراضي)
        self.entry_date_label = ttk.Label(form_frame, text="تاريخ الدخول:")
        self.entry_date_entry = DateEntry(form_frame, textvariable=self.entry_date, date_pattern="yyyy-mm-dd")
        self.exit_date_label = ttk.Label(form_frame, text="تاريخ الخروج:")
        self.exit_date_entry = DateEntry(form_frame, textvariable=self.exit_date, date_pattern="yyyy-mm-dd")

        # حقل "المبلغ المتبقي" (مخفي بشكل افتراضي)
        self.remaining_amount_label = ttk.Label(form_frame, text="المبلغ المتبقي أقل من:")
        self.remaining_amount_entry = ttk.Entry(form_frame, textvariable=self.remaining_amount_threshold, width=30)

        # زر التصدير إلى Excel
        export_excel_button = ttk.Button(form_frame, text="تصدير إلى Excel", command=self.export_to_excel)
        export_excel_button.grid(row=6, column=0, columnspan=2, pady=10, padx=5, sticky="ew")

    def toggle_fields(self, event=None):
        """
        إظهار أو إخفاء حقول التاريخ وحقل "المبلغ المتبقي" بناءً على الخيار المحدد.
        """
        if self.export_option.get() == "حسب تاريخ الدخول":
            self.entry_date_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
            self.entry_date_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
            self.exit_date_label.grid_remove()
            self.exit_date_entry.grid_remove()
            self.remaining_amount_label.grid_remove()
            self.remaining_amount_entry.grid_remove()
        elif self.export_option.get() == "حسب تاريخ الخروج":
            self.exit_date_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")
            self.exit_date_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")
            self.entry_date_label.grid_remove()
            self.entry_date_entry.grid_remove()
            self.remaining_amount_label.grid_remove()
            self.remaining_amount_entry.grid_remove()
        elif self.export_option.get() == "بيانات بها مبالغ متبقية":
            self.remaining_amount_label.grid(row=4, column=0, padx=5, pady=5, sticky="e")
            self.remaining_amount_entry.grid(row=4, column=1, padx=5, pady=5, sticky="w")
            self.entry_date_label.grid_remove()
            self.entry_date_entry.grid_remove()
            self.exit_date_label.grid_remove()
            self.exit_date_entry.grid_remove()
        else:
            self.entry_date_label.grid_remove()
            self.entry_date_entry.grid_remove()
            self.exit_date_label.grid_remove()
            self.exit_date_entry.grid_remove()
            self.remaining_amount_label.grid_remove()
            self.remaining_amount_entry.grid_remove()

    def get_filtered_data(self):
        """
        جلب البيانات المصفاة بناءً على الخيارات المحددة.
        """
        query = f"SELECT * FROM {self.table_name}"
        conditions = []

        if self.export_option.get() == "حسب تاريخ الدخول":
            if not self.entry_date.get():
                messagebox.showwarning("تحذير", "يجب تحديد تاريخ الدخول.")
                return None
            conditions.append(f"entry_date = '{self.entry_date.get()}'")

        elif self.export_option.get() == "حسب تاريخ الخروج":
            if not self.exit_date.get():
                messagebox.showwarning("تحذير", "يجب تحديد تاريخ الخروج.")
                return None
            conditions.append(f"exit_date = '{self.exit_date.get()}'")

        elif self.export_option.get() == "بيانات بها مبالغ متبقية":
            conditions.append(f"remaining_amount > {self.remaining_amount_threshold.get()}")

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        # جلب البيانات من قاعدة البيانات
        data = self.db_manager.execute_read_query(query)

        if not data:
            messagebox.showinfo("معلومات", "لا توجد بيانات للتصدير.")
            return None

        return data

    def format_currency(self, currency_code):
        """
        تحويل رمز العملة المخزن في قاعدة البيانات إلى نص.
        """
        currency_map = {"1": "ر.ي", "2": "ر.س", "3": "دولار"}
        return currency_map.get(str(currency_code), "ر.ي")  # افتراضيًا ر.ي إذا لم يتم العثور على الرمز

    def merge_and_remove_currency(self, df):
        """
        دمج العملة مع الأعمدة المالية وحذف عمود العملة.
        """
        # تحويل رمز العملة إلى نص باستخدام دالة format_currency
        df["العملة"] = df["العملة"].apply(self.format_currency)

        # دمج العملة مع الأعمدة المالية
        df["التكلفة"] = df["التكلفة"].astype(str) + " " + df["العملة"]
        df["المبلغ المدفوع"] = df["المبلغ المدفوع"].astype(str) + " " + df["العملة"]
        df["المبلغ المتبقي"] = df["المبلغ المتبقي"].astype(str) + " " + df["العملة"]

        # حذف عمود العملة
        df.drop(columns=["العملة"], inplace=True)

        return df

    def apply_rtl_to_excel(self, file_path):
        """
        تطبيق توجيه النص من اليمين إلى اليسار على ملف Excel.
        """
        # تحميل ملف Excel
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # تطبيق توجيه النص RTL على جميع الخلايا
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="right", vertical="center")

        # حفظ التغييرات
        workbook.save(file_path)

    def apply_colors_to_excel(self, file_path):
        """
        تطبيق الألوان على الصفوف والأعمدة في ملف Excel.
        """
        # تحميل ملف Excel
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # تعريف الألوان
        gray_100 = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # رمادي فاتح
        white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # أبيض

        # تطبيق الألوان على الصفوف
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # تجاهل الصف الأول (العناوين)
            if row_idx % 2 == 0:  # الصفوف الزوجية
                for cell in row:
                    cell.fill = gray_100
            else:  # الصفوف الفردية
                for cell in row:
                    cell.fill = white

        # حفظ التغييرات
        workbook.save(file_path)

    def export_to_excel(self):
        """
        تصدير البيانات إلى ملف Excel.
        """
        data = self.get_filtered_data()
        if not data:
            return

        # تحويل البيانات إلى DataFrame
        columns = [description[0] for description in self.db_manager.cursor.description]
        df = pd.DataFrame(data, columns=columns)

        # تحويل أسماء الأعمدة إلى العربية
        arabic_columns = {
            "id": "الرقم",
            "name": "الاسم",
            "passport_number": "رقم الجواز",
            "phone_number": "رقم الهاتف",
            "sponsor_name": "اسم الكفيل",
            "sponsor_number": "رقم الكفيل",
            "cost": "التكلفة",
            "paid": "المبلغ المدفوع",
            "remaining_amount": "المبلغ المتبقي",
            "entry_date": "تاريخ الدخول",
            "exit_date": "تاريخ الخروج",
            "status": "الحالة",
            "currency": "العملة"
        }
        df.rename(columns=arabic_columns, inplace=True)

        # دمج العملة مع الأعمدة المالية وحذف عمود العملة
        df = self.merge_and_remove_currency(df)

        # تحديد مسار حفظ الملف
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        file_path = os.path.join(downloads_path, f"{self.filename.get()}.xlsx")

        # تصدير البيانات إلى Excel
        df.to_excel(file_path, index=False)

        # تطبيق توجيه النص RTL على ملف Excel
        self.apply_rtl_to_excel(file_path)

        # تطبيق الألوان على ملف Excel
        self.apply_colors_to_excel(file_path)

        messagebox.showinfo("نجاح", f"تم تصدير البيانات بنجاح إلى: {file_path}")

        # إغلاق النافذة بعد التصدير
        self.export_window.destroy()